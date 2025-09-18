import os
import sys
import uuid
import shutil
from datetime import datetime
from pathlib import Path

from flask import (
    Flask, render_template, request, redirect, url_for,
    flash, jsonify, send_file
)
from werkzeug.utils import secure_filename

# QR
import qrcode
import qrcode.image.svg

# Excel
import openpyxl

# --- SQLAlchemy (DB-agnostic: Postgres in prod, SQLite locally) ---
from sqlalchemy import (
    create_engine, MetaData, Table, Column, Integer, String, Text, Float,
    DateTime, Boolean, select, func, insert, update, delete, and_, or_
)
from sqlalchemy.engine import Engine
from sqlalchemy.exc import IntegrityError

# -------------------------------------------------------------------
# Paths & Flask (persistent storage + SPA layout)
# -------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"
TEMPLATES_DIR = BASE_DIR / "templates"
STATIC_DIR.mkdir(parents=True, exist_ok=True)

# Persistent disk (Render: set DATA_DIR=/opt/render/project/src/data)
DATA_DIR = Path(os.environ.get("DATA_DIR", BASE_DIR / "data"))
DATA_DIR.mkdir(parents=True, exist_ok=True)

# --- DB URL normalization (Render) ---------------------------------
def normalize_db_url():
    db_url = (
        os.environ.get("DATABASE_URL")
        or os.environ.get("POSTGRES_URL")
        or os.environ.get("DATABASE_INTERNAL_URL")
        or ""
    ).strip()
    if not db_url:
        return ""
    if db_url.startswith("postgres://"):
        db_url = db_url.replace("postgres://", "postgresql://", 1)
    if "sslmode=" not in db_url:
        db_url += ("&" if "?" in db_url else "?") + "sslmode=require"
    return db_url

DATABASE_URL = normalize_db_url()
if DATABASE_URL:
    ENGINE_URL = DATABASE_URL
else:
    ENGINE_URL = f"sqlite:///{(DATA_DIR / 'attendance.db').as_posix()}"

# Flask
app = Flask(__name__, static_folder=str(STATIC_DIR), template_folder=str(TEMPLATES_DIR))
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "dev-dev-dev")

# uploads
MAX_UPLOAD_MB = 5
app.config["MAX_CONTENT_LENGTH"] = MAX_UPLOAD_MB * 1024 * 1024
ALLOWED_XLSX_EXT = {".xlsx"}

# Persistent media dirs (live on the mounted disk)
MEDIA_QR_DIR = DATA_DIR / "qrcodes"
UPLOADS_DIR = DATA_DIR / "uploads"
MEDIA_QR_DIR.mkdir(parents=True, exist_ok=True)
UPLOADS_DIR.mkdir(parents=True, exist_ok=True)

# Robust symlink into /static so existing paths continue to work
def _ensure_symlink(target: Path, link: Path):
    """
    Ensure 'link' is a symlink to 'target'.
    If 'link' exists as a real directory, migrate its contents into 'target',
    remove it, then create the symlink. Fixes broken images after redeploys.
    """
    try:
        link.parent.mkdir(parents=True, exist_ok=True)
        target.mkdir(parents=True, exist_ok=True)

        if link.exists() and link.is_symlink():
            return

        if link.exists() and link.is_dir() and not link.is_symlink():
            for child in link.iterdir():
                dest = target / child.name
                shutil.move(str(child), str(dest))
            shutil.rmtree(link)

        if not link.exists():
            link.symlink_to(target, target_is_directory=True)
    except Exception as e:
        app.logger.warning("Could not ensure symlink %s -> %s: %s", link, target, e)

_ensure_symlink(MEDIA_QR_DIR, STATIC_DIR / "qrcodes")
_ensure_symlink(UPLOADS_DIR, STATIC_DIR / "uploads")

# Where to actually write QR files
QR_DIR = MEDIA_QR_DIR

# -------------------------------------------------------------------
# Database (SQLAlchemy Core)
# -------------------------------------------------------------------
engine: Engine = create_engine(ENGINE_URL, pool_pre_ping=True, future=True)
metadata = MetaData()

# Tables (portable across Postgres & SQLite)
workers = Table(
    "workers", metadata,
    Column("id", Integer, primary_key=True),
    Column("name", String(255), nullable=False),
    Column("token_id", String(255), nullable=False, unique=True),
    Column("department", String(255), nullable=False),
    Column("line", String(255)),
    # NOTE: no server_default here; set by PG bootstrap DDL
    Column("active", Boolean, nullable=False),
    Column("qrcode_path", String(512)),
    Column("qrcode_svg_path", String(512)),
    Column("created_at", DateTime(timezone=True), server_default=func.now()),
    Column("updated_at", DateTime(timezone=True), server_default=func.now(), onupdate=func.now()),
)

operations = Table(
    "operations", metadata,
    Column("id", Integer, primary_key=True),
    Column("seq_no", Integer),
    Column("op_no", String(100)),
    Column("description", Text),
    Column("machine", String(100)),
    Column("department", String(255)),
    Column("std_min", Float),
    Column("piece_rate", Float),
    Column("created_at", DateTime(timezone=True), server_default=func.now()),
)

bundles = Table(
    "bundles", metadata,
    Column("id", Integer, primary_key=True),
    Column("bundle_no", String(255)),
    Column("order_no", String(255)),
    Column("style", String(255)),
    Column("color", String(255)),
    Column("size", String(50)),
    Column("quantity", Integer),
    Column("status", String(50), server_default="Pending"),
    Column("created_at", DateTime(timezone=True), server_default=func.now()),
)

production_orders = Table(
    "production_orders", metadata,
    Column("id", Integer, primary_key=True),
    Column("order_no", String(255), unique=True),
    Column("style", String(255)),
    Column("quantity", Integer),
    Column("buyer", String(255)),
    Column("created_at", DateTime(timezone=True), server_default=func.now()),
)

scans = Table(
    "scans", metadata,
    Column("id", Integer, primary_key=True),
    Column("code", String(255), nullable=False),
    Column("worker_id", Integer),
    Column("bundle_id", Integer),
    Column("created_at", DateTime(timezone=True), server_default=func.now()),
)

file_uploads = Table(
    "file_uploads", metadata,
    Column("id", Integer, primary_key=True),
    Column("filename", String(255), nullable=False),
    Column("original_filename", String(255), nullable=False),
    Column("file_type", String(50), nullable=False),
    Column("file_path", String(1024), nullable=False),
    Column("uploaded_at", DateTime(timezone=True), server_default=func.now()),
)

# --- Boot-time schema fixer (handles missing columns on Postgres) ---
def ensure_pg_schema():
    # Only run against Postgres
    if not DATABASE_URL or not DATABASE_URL.startswith("postgresql://"):
        print("Schema bootstrap: using SQLite or no DATABASE_URL; skipping.", file=sys.stderr)
        return
    ddl = """
    -- Ensure minimal tables exist
    CREATE TABLE IF NOT EXISTS workers (
      id SERIAL PRIMARY KEY,
      name text,
      token_id text UNIQUE,
      department text,
      line text
    );
    CREATE TABLE IF NOT EXISTS operations (
      id SERIAL PRIMARY KEY,
      seq_no integer,
      op_no text,
      description text,
      machine text,
      department text,
      std_min double precision,
      piece_rate double precision,
      created_at timestamptz DEFAULT now()
    );
    CREATE TABLE IF NOT EXISTS bundles ( id SERIAL PRIMARY KEY );
    CREATE TABLE IF NOT EXISTS production_orders (
      id SERIAL PRIMARY KEY,
      order_no text UNIQUE,
      style text,
      quantity integer,
      buyer text,
      created_at timestamptz DEFAULT now()
    );
    CREATE TABLE IF NOT EXISTS scans ( id SERIAL PRIMARY KEY );
    CREATE TABLE IF NOT EXISTS file_uploads (
      id SERIAL PRIMARY KEY,
      filename text NOT NULL,
      original_filename text NOT NULL,
      file_type text NOT NULL,
      file_path text NOT NULL,
      uploaded_at timestamptz DEFAULT now()
    );

    -- workers: expected columns
    ALTER TABLE workers ADD COLUMN IF NOT EXISTS active boolean NOT NULL DEFAULT true;
    ALTER TABLE workers ADD COLUMN IF NOT EXISTS qrcode_path text;
    ALTER TABLE workers ADD COLUMN IF NOT EXISTS qrcode_svg_path text;
    ALTER TABLE workers ADD COLUMN IF NOT EXISTS created_at timestamptz DEFAULT now();
    ALTER TABLE workers ADD COLUMN IF NOT EXISTS updated_at timestamptz DEFAULT now();

    -- bundles: rename bundle_code -> bundle_no if present
    DO $$
    BEGIN
      IF EXISTS (
        SELECT 1 FROM information_schema.columns
        WHERE table_name='bundles' AND column_name='bundle_code'
      ) AND NOT EXISTS (
        SELECT 1 FROM information_schema.columns
        WHERE table_name='bundles' AND column_name='bundle_no'
      ) THEN
        ALTER TABLE bundles RENAME COLUMN bundle_code TO bundle_no;
      END IF;
    END $$;

    -- bundles: expected columns
    ALTER TABLE bundles ADD COLUMN IF NOT EXISTS bundle_no text;
    ALTER TABLE bundles ADD COLUMN IF NOT EXISTS order_no text;
    ALTER TABLE bundles ADD COLUMN IF NOT EXISTS style text;
    ALTER TABLE bundles ADD COLUMN IF NOT EXISTS color text;
    ALTER TABLE bundles ADD COLUMN IF NOT EXISTS size text;
    ALTER TABLE bundles ADD COLUMN IF NOT EXISTS quantity integer;
    ALTER TABLE bundles ADD COLUMN IF NOT EXISTS status text DEFAULT 'Pending';
    ALTER TABLE bundles ADD COLUMN IF NOT EXISTS created_at timestamptz DEFAULT now();

    -- scans: expected columns
    ALTER TABLE scans ADD COLUMN IF NOT EXISTS code text;
    ALTER TABLE scans ADD COLUMN IF NOT EXISTS worker_id integer;
    ALTER TABLE scans ADD COLUMN IF NOT EXISTS bundle_id integer;
    ALTER TABLE scans ADD COLUMN IF NOT EXISTS created_at timestamptz DEFAULT now();
    """
    try:
        with engine.begin() as conn:
            conn.exec_driver_sql(ddl)
        print("Schema bootstrap: ensured ✔", file=sys.stderr)
    except Exception as e:
        print(f"Schema bootstrap failed: {e}", file=sys.stderr)

def init_db():
    ensure_pg_schema()
    metadata.create_all(engine)

# -------------------------------------------------------------------
# Helpers
# -------------------------------------------------------------------
def ci_like(column, term: str):
    """Portable case-insensitive LIKE for Postgres/SQLite. Pass a pattern like '%foo%'."""
    return func.lower(column).like(term.lower())

def fmt_ts(v):
    try:
        return v.isoformat()
    except Exception:
        return str(v) if v else ""

# -------------------------------------------------------------------
# QR helpers (SVG + PNG) + self-healing
# -------------------------------------------------------------------
def generate_qr_files(token_id: str, worker_id: int) -> tuple[str, str]:
    """
    Create PNG + SVG QR for token_id under persistent qrcodes/.
    Returns (png_rel_path, svg_rel_path) relative to /static (qrcodes/...).
    """
    _ensure_symlink(MEDIA_QR_DIR, STATIC_DIR / "qrcodes")

    ts = datetime.now().strftime("%Y%m%d%H%M%S")
    base = f"qrcode_{token_id}_{worker_id}_{ts}"
    png_path = QR_DIR / f"{base}.png"
    svg_path = QR_DIR / f"{base}.svg"

    # SVG
    svg_factory = qrcode.image.svg.SvgImage
    svg_img = qrcode.make(token_id, image_factory=svg_factory)
    svg_path.parent.mkdir(parents=True, exist_ok=True)
    svg_img.save(str(svg_path))

    # PNG
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(token_id)
    qr.make(fit=True)
    png_img = qr.make_image(fill_color="black", back_color="white")
    png_path.parent.mkdir(parents=True, exist_ok=True)
    png_img.save(str(png_path), format="PNG")

    return f"qrcodes/{png_path.name}", f"qrcodes/{svg_path.name}"

def _ensure_qr_present(conn, worker_row: dict) -> tuple[str, str]:
    """
    If the worker's QR PNG/SVG is missing on disk (e.g., after a deploy),
    regenerate and update DB. Return (png_rel, svg_rel).
    """
    _ensure_symlink(MEDIA_QR_DIR, STATIC_DIR / "qrcodes")

    png_rel = worker_row.get("qrcode_path")
    svg_rel = worker_row.get("qrcode_svg_path")
    needs_regen = True

    if png_rel:
        p = STATIC_DIR / png_rel
        if p.exists():
            needs_regen = False

    if needs_regen:
        new_png_rel, new_svg_rel = generate_qr_files(worker_row["token_id"], worker_row["id"])
        conn.execute(
            update(workers)
            .where(workers.c.id == worker_row["id"])
            .values(qrcode_path=new_png_rel, qrcode_svg_path=new_svg_rel, updated_at=func.now())
        )
        return new_png_rel, new_svg_rel
    else:
        return png_rel or "", svg_rel or ""

def delete_qr_files(qr_png_rel: str | None, qr_svg_rel: str | None):
    for rel in (qr_png_rel, qr_svg_rel):
        if not rel:
            continue
        p = STATIC_DIR / rel
        try:
            if p.exists():
                p.unlink()
        except Exception as e:
            app.logger.error("Failed to delete QR file %s: %s", p, e)

# -------------------------------------------------------------------
# UI (SPA index + deep-link helpers)
# -------------------------------------------------------------------
@app.get("/")
def index():
    with engine.begin() as conn:
        rows = conn.execute(
            select(workers).order_by(workers.c.created_at.desc())
        ).mappings().all()

        # Self-heal missing QR files before rendering
        fixed = []
        for r in rows:
            rd = dict(r)
            png_rel, svg_rel = _ensure_qr_present(conn, rd)
            rd["qrcode_path"] = png_rel
            rd["qrcode_svg_path"] = svg_rel
            fixed.append(rd)

    return render_template("index.html", workers=fixed)

@app.get("/dashboard")
def dashboard():
    return redirect(url_for("index"))

def _section_redirect(section: str):
    return redirect(url_for("index") + f"#{section}")

@app.get("/workers")
def workers_page():
    return _section_redirect("workers")

@app.get("/operations")
def operations_page():
    return _section_redirect("operations")

@app.get("/bundles")
def bundles_page():
    return _section_redirect("bundles")

@app.get("/production-order")
def production_order_page():
    return _section_redirect("production-order")

@app.get("/file-upload")
def file_upload_page():
    return _section_redirect("file-upload")

@app.get("/scanner")
def scanner_page():
    return _section_redirect("scanner")

@app.get("/reports")
def reports_page():
    return _section_redirect("reports")

# -------------------------------------------------------------------
# API endpoints
# -------------------------------------------------------------------
@app.get("/api/dashboard-stats")
def api_dashboard_stats():
    try:
        with engine.begin() as conn:
            active_workers = conn.execute(
                select(func.count()).select_from(workers).where(workers.c.active.is_(True))
            ).scalar_one()
            total_bundles = conn.execute(select(func.count()).select_from(bundles)).scalar_one()
            total_operations = conn.execute(select(func.count()).select_from(operations)).scalar_one()
            total_earnings = conn.execute(
                select(func.coalesce(func.sum(operations.c.piece_rate * 5), 0.0))
            ).scalar_one() or 0.0

        return jsonify({
            "activeWorkers": int(active_workers or 0),
            "totalBundles": int(total_bundles or 0),
            "totalOperations": int(total_operations or 0),
            "totalEarnings": float(total_earnings or 0.0)
        })
    except Exception as e:
        app.logger.error("dashboard-stats error: %s", e)
        return jsonify({"activeWorkers": 0, "totalBundles": 0, "totalOperations": 0, "totalEarnings": 0})

@app.get("/api/chart-data")
def api_chart_data():
    try:
        with engine.begin() as conn:
            bs = conn.execute(
                select(bundles.c.status, func.count().label("c")).group_by(bundles.c.status)
            ).all()
            bundle_status = {r[0]: r[1] for r in bs}

            dw = conn.execute(
                select(workers.c.department, func.count().label("c")).group_by(workers.c.department)
            ).all()
            dept = {(r[0] or "Unknown"): r[1] for r in dw}

        return jsonify({"bundleStatus": bundle_status, "departmentWorkload": dept})
    except Exception as e:
        app.logger.error("chart-data error: %s", e)
        return jsonify({"bundleStatus": {}, "departmentWorkload": {}})

@app.get("/api/recent-activity")
def api_recent_activity():
    try:
        with engine.begin() as conn:
            rows = conn.execute(
                select(scans.c.code, scans.c.created_at)
                .order_by(scans.c.created_at.desc())
                .limit(10)
            ).mappings().all()

        data = [{
            "type": "Scan",
            "description": r["code"],
            "created_at": fmt_ts(r["created_at"])
        } for r in rows]
        return jsonify(data)
    except Exception as e:
        app.logger.error("recent-activity error: %s", e)
        return jsonify([])

@app.get("/api/operations")
def api_operations():
    search = (request.args.get("search") or "").strip()
    stmt = select(operations)
    if search:
        like = f"%{search}%"
        stmt = stmt.where(or_(
            ci_like(operations.c.description, like),
            ci_like(operations.c.op_no, like)
        ))
    stmt = stmt.order_by(func.coalesce(operations.c.seq_no, 999999), operations.c.id)

    with engine.begin() as conn:
        rows = conn.execute(stmt).mappings().all()
    return jsonify([dict(r) for r in rows])

@app.get("/api/bundles")
def api_bundles():
    with engine.begin() as conn:
        rows = conn.execute(select(bundles).order_by(bundles.c.created_at.desc())).mappings().all()
    return jsonify([dict(r) for r in rows])

@app.get("/api/production-order")
def api_production_order():
    with engine.begin() as conn:
        row = conn.execute(
            select(production_orders).order_by(production_orders.c.created_at.desc()).limit(1)
        ).mappings().first()
    return jsonify(dict(row) if row else {})

@app.get("/api/workers")
def api_workers():
    search = (request.args.get("search") or "").strip()
    department = (request.args.get("department") or "").strip()
    status = (request.args.get("status") or "").strip()  # "Active" or "Idle" or ""

    stmt = select(workers)
    conds = []
    if search:
        like = f"%{search}%"
        conds.append(or_(
            ci_like(workers.c.name, like),
            ci_like(workers.c.token_id, like),
            ci_like(workers.c.department, like),
            ci_like(workers.c.line, like)
        ))
    if department:
        conds.append(workers.c.department == department)
    if status:
        if status.lower() == "active":
            conds.append(workers.c.active.is_(True))
        elif status.lower() == "idle":
            conds.append(workers.c.active.is_(False))

    if conds:
        stmt = stmt.where(and_(*conds))
    stmt = stmt.order_by(workers.c.created_at.desc())

    with engine.begin() as conn:
        rows = conn.execute(stmt).mappings().all()

        out = []
        for r in rows:
            rd = dict(r)
            # Self-heal QR before returning to the UI
            png_rel, svg_rel = _ensure_qr_present(conn, rd)
            rd["qrcode_path"] = png_rel
            rd["qrcode_svg_path"] = svg_rel

            out.append({
                "id": rd["id"],
                "name": rd["name"],
                "token_id": rd["token_id"],
                "department": rd["department"],
                "line": rd["line"],
                "active": bool(rd["active"]),
                "qrcode_path": rd["qrcode_path"],
                "created_at": fmt_ts(rd["created_at"]),
                "updated_at": fmt_ts(rd["updated_at"]),
            })

    return jsonify(out)

# -------------------------------------------------------------------
# Single worker add / edit / delete / download QR
# -------------------------------------------------------------------
@app.route("/add", methods=["GET", "POST"])
def add_worker():
    if request.method == "GET":
        return render_template("add_worker.html")

    name = (request.form.get("name") or "").strip()
    token_id = (request.form.get("token_id") or "").strip()
    department = (request.form.get("department") or "").strip()
    line = (request.form.get("line") or "").strip()
    active_val = request.form.get("active")
    active_bool = True if active_val in ("1", "on", "true", "True") else False

    if not token_id:
        flash("Token ID cannot be empty.", "error")
        return redirect(url_for("add_worker"))

    try:
        with engine.begin() as conn:
            res = conn.execute(insert(workers).values(
                name=name,
                token_id=token_id,
                department=department,
                line=line,
                active=active_bool,
            ))
            worker_id = res.inserted_primary_key[0]

            png_rel, svg_rel = generate_qr_files(token_id, worker_id)
            conn.execute(update(workers).where(workers.c.id == worker_id).values(
                qrcode_path=png_rel,
                qrcode_svg_path=svg_rel,
                updated_at=func.now()
            ))

        flash("Worker added successfully!", "success")
        return redirect(url_for("index"))
    except IntegrityError:
        flash("Token ID already exists. Please use a unique token.", "error")
        return redirect(url_for("add_worker"))
    except Exception as e:
        app.logger.error("add_worker error: %s", e)
        flash("Server error while adding worker.", "error")
        return redirect(url_for("add_worker"))

@app.route("/edit/<int:worker_id>", methods=["GET", "POST"])
def edit_worker(worker_id: int):
    with engine.begin() as conn:
        worker_row = conn.execute(select(workers).where(workers.c.id == worker_id)).mappings().first()
        if not worker_row:
            flash("Worker not found.", "error")
            return redirect(url_for("index"))

    if request.method == "GET":
        return render_template("edit_worker.html", worker=worker_row)

    name = (request.form.get("name") or "").strip()
    department = (request.form.get("department") or "").strip()
    line = (request.form.get("line") or "").strip()
    active_val = request.form.get("active")
    active_bool = True if active_val in ("1", "on", "true", "True") else False

    try:
        with engine.begin() as conn:
            conn.execute(update(workers)
                .where(workers.c.id == worker_id)
                .values(
                    name=name,
                    department=department,
                    line=line,
                    active=active_bool,
                    updated_at=func.now()
                )
            )
        flash("Worker updated successfully!", "success")
    except Exception as e:
        app.logger.error("edit_worker error: %s", e)
        flash("Server error while updating worker.", "error")
    return redirect(url_for("index"))

@app.route("/delete/<int:worker_id>", methods=["POST", "GET"])
def delete_worker(worker_id: int):
    try:
        with engine.begin() as conn:
            row = conn.execute(
                select(workers.c.qrcode_path, workers.c.qrcode_svg_path)
                .where(workers.c.id == worker_id)
            ).first()
            if not row:
                flash("Worker not found.", "error")
                return redirect(url_for("index"))
            delete_qr_files(row[0], row[1])
            conn.execute(delete(workers).where(workers.c.id == worker_id))

        flash("Worker deleted.", "success")
    except Exception as e:
        app.logger.error("delete_worker error: %s", e)
        flash("Server error while deleting worker.", "error")
    return redirect(url_for("index"))

@app.get("/download_qr/<int:worker_id>")
def download_qr(worker_id: int):
    with engine.begin() as conn:
        row = conn.execute(
            select(workers.c.token_id, workers.c.qrcode_path).where(workers.c.id == worker_id)
        ).first()
        if not row or not row[1]:
            flash("QR not available.", "error")
            return redirect(url_for("index"))

    p = STATIC_DIR / row[1]
    if not p.exists():
        # self-heal if someone tries to download right after a deploy
        with engine.begin() as conn:
            png_rel, _ = generate_qr_files(row[0], worker_id)
            conn.execute(
                update(workers).where(workers.c.id == worker_id).values(qrcode_path=png_rel, updated_at=func.now())
            )
        p = STATIC_DIR / png_rel

    if not p.exists():
        flash("QR file missing on disk.", "error")
        return redirect(url_for("index"))
    return send_file(str(p), mimetype="image/png", as_attachment=True, download_name=f"qr_{row[0]}.png")

# -------------------------------------------------------------------
# NEW: Print pages (single & batch)
# -------------------------------------------------------------------
@app.get("/print_qr/<int:worker_id>")
def print_qr(worker_id: int):
    with engine.begin() as conn:
        r = conn.execute(select(workers).where(workers.c.id == worker_id)).mappings().first()
        if not r:
            flash("Worker not found.", "error")
            return redirect(url_for("index"))
        rd = dict(r)
        png_rel, svg_rel = _ensure_qr_present(conn, rd)
        rd["qrcode_path"] = png_rel
        rd["qrcode_svg_path"] = svg_rel
    return render_template("print_qr.html", worker=rd)

@app.get("/print_qrs")
def print_qrs():
    ids_param = (request.args.get("ids") or "").strip()
    if not ids_param:
        flash("No workers selected to print.", "error")
        return redirect(url_for("index"))
    try:
        ids = [int(x) for x in ids_param.split(",") if x.strip().isdigit()]
    except Exception:
        flash("Invalid ids.", "error")
        return redirect(url_for("index"))
    if not ids:
        flash("No valid ids provided.", "error")
        return redirect(url_for("index"))

    items = []
    with engine.begin() as conn:
        rows = conn.execute(select(workers).where(workers.c.id.in_(ids)).order_by(workers.c.id)).mappings().all()
        for r in rows:
            rd = dict(r)
            png_rel, svg_rel = _ensure_qr_present(conn, rd)
            rd["qrcode_path"] = png_rel
            rd["qrcode_svg_path"] = svg_rel
            items.append(rd)

    return render_template("print_qrs.html", workers=items)

# -------------------------------------------------------------------
# NEW: Bulk delete API
# -------------------------------------------------------------------
@app.post("/api/workers/delete")
def api_workers_delete():
    payload = request.get_json(silent=True) or {}
    ids = payload.get("ids", [])
    if not isinstance(ids, list) or not all(isinstance(i, int) for i in ids):
        return jsonify({"error": "Provide JSON { ids: [int,...] }"}), 400
    if not ids:
        return jsonify({"deleted": 0})

    try:
        with engine.begin() as conn:
            rows = conn.execute(
                select(workers.c.id, workers.c.qrcode_path, workers.c.qrcode_svg_path)
                .where(workers.c.id.in_(ids))
            ).all()
            for _id, png_rel, svg_rel in rows:
                delete_qr_files(png_rel, svg_rel)

            result = conn.execute(delete(workers).where(workers.c.id.in_(ids)))
            deleted_count = result.rowcount or 0
        return jsonify({"deleted": int(deleted_count)})
    except Exception as e:
        app.logger.error("bulk delete error: %s", e)
        return jsonify({"error": "Server error during delete"}), 500

# -------------------------------------------------------------------
# Bulk Excel upload with dedupe
#  (explicit endpoint so url_for('upload_workers') is always found)
# -------------------------------------------------------------------
@app.route("/upload_workers", methods=["POST"], endpoint="upload_workers")
@app.route("/upload_excel", methods=["POST"])
def upload_workers():
    f = request.files.get("file")
    if not f or not f.filename:
        flash("No file selected.", "error")
        return redirect(url_for("index"))

    ext = Path(f.filename).suffix.lower()
    if ext not in ALLOWED_XLSX_EXT:
        flash("Invalid file. Please upload a .xlsx file.", "error")
        return redirect(url_for("index"))

    temp_path = UPLOADS_DIR / f"{uuid.uuid4()}_{secure_filename(f.filename)}"
    f.save(temp_path)

    added = skipped = invalid = 0
    skipped_tokens: list[str] = []

    try:
        wb = openpyxl.load_workbook(temp_path)
        ws = wb.active

        header = [str(c).strip().lower() if c is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        required = ["name", "token_id", "department", "line", "active"]
        missing = [h for h in required if h not in header]
        if missing:
            flash(f"Excel missing required headers: {', '.join(missing)}", "error")
            temp_path.unlink(missing_ok=True)
            return redirect(url_for("index"))

        idx = {h: header.index(h) for h in header}

        with engine.begin() as conn:
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    name = (str(row[idx["name"]]).strip() if row[idx["name"]] is not None else "")
                    token_id = (str(row[idx["token_id"]]).strip() if row[idx["token_id"]] is not None else "")
                    department = (str(row[idx["department"]]).strip() if row[idx["department"]] is not None else "")
                    line = (str(row[idx["line"]]).strip() if row[idx["line"]] is not None else "")
                    active_cell = row[idx["active"]]
                    active_bool = str(active_cell).strip().lower() in ("1", "true", "yes", "y")
                except Exception:
                    invalid += 1
                    continue

                if not token_id:
                    invalid += 1
                    continue

                exists = conn.execute(
                    select(workers.c.id).where(workers.c.token_id == token_id)
                ).first()
                if exists:
                    skipped += 1
                    if len(skipped_tokens) < 10:
                        skipped_tokens.append(token_id)
                    continue

                res = conn.execute(insert(workers).values(
                    name=name,
                    token_id=token_id,
                    department=department,
                    line=line,
                    active=active_bool,
                ))
                worker_id = res.inserted_primary_key[0]

                png_rel, svg_rel = generate_qr_files(token_id, worker_id)
                conn.execute(update(workers).where(workers.c.id == worker_id).values(
                    qrcode_path=png_rel,
                    qrcode_svg_path=svg_rel,
                    updated_at=func.now()
                ))
                added += 1
    except Exception as e:
        app.logger.error("Excel processing error: %s", e)
        flash("Error processing Excel file.", "error")
        try:
            temp_path.unlink(missing_ok=True)
        except Exception:
            pass
        return redirect(url_for("index"))

    try:
        temp_path.unlink(missing_ok=True)
    except Exception:
        pass

    summary = f"Upload complete. Added: {added}, Skipped (duplicates): {skipped}, Invalid: {invalid}"
    if skipped_tokens:
        summary += f" | Skipped token_ids (first 10): {', '.join(skipped_tokens)}"
    flash(summary, "success")
    return redirect(url_for("index"))

# -------------------------------------------------------------------
# Health
# -------------------------------------------------------------------
@app.get("/health")
def health():
    _ensure_symlink(MEDIA_QR_DIR, STATIC_DIR / "qrcodes")
    return "OK", 200

# -------------------------------------------------------------------
# Startup
# -------------------------------------------------------------------
if __name__ == "__main__":
    init_db()
    _ensure_symlink(MEDIA_QR_DIR, STATIC_DIR / "qrcodes")
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", "5000")))
else:
    init_db()
    _ensure_symlink(MEDIA_QR_DIR, STATIC_DIR / "qrcodes")
